import pandas as pd
import os
import sys
import logging
from datetime import datetime
import requests

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../src')))
from utils import normalize_nit
from dian_processor import check_nit_dian
from bdme_scraper import consult_batch_bdme

# Configuración del logger
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def check_facturador_electronico(nit: str) -> bool:
    """
    [Simulado] Consulta la API de facturadores electrónicos de la DIAN.
    Retorna True si el NIT está habilitado para facturar, False en caso contrario o error.
    Para el MVP construiremos un mock basándonos en si el NIT termina en un digito par (simulando casos variados) 
    ya que no hay una API publica abierta sin autenticación institucional provista en el prompt.
    """
    nit_norm = normalize_nit(nit)
    try:
        # En producción aqui se haría un requests.get/post a la API oficial de facturación
        # simularemos una demora de red y un resultado heuristico determinístico
        logger.info(f"Consultando estado de facturador electrónico para NIT {nit_norm}")
        # Simulando resultado basado en NIT
        if int(nit_norm[-1]) % 2 == 0 or "900" in nit_norm:
            return True
        else:
             return False # simulamos no habilitados los impares para test de reglas
    except Exception as e:
        logger.error(f"Error consultando API facturador para {nit_norm}: {e}")
        return False

def calculate_risk(en_ficticios: bool, estado_bdme: str, facturador_habilitado: bool) -> tuple[str, str]:
    """
    Calcula el nivel de riesgo y la recomendación dada la matriz de reglas de precedencia:
    - NIT en Proveedores Ficticios (DIAN) -> ALTO
    - NIT en BDME como deudor moroso o no habilitado facturador -> MEDIO
    - Consulta BDME falló (INDETERMINADO) -> REVISAR
    - Ninguna alerta -> BAJO
    Retorna la tupla (Nivel, Recomendacion)
    """
    if en_ficticios:
        return "ALTO", "Bloquear pago + alerta inmediata"
    
    if estado_bdme == "EN_MORA" or not facturador_habilitado:
        return "MEDIO", "Revisar con asesor tributario antes de pagar"
        
    if estado_bdme == "INDETERMINADO":
        return "REVISAR", "Falla de consulta externa, verificar manualmente"
        
    return "BAJO", "Proveedor verificado — sin alertas"

def get_audit_data(client_file_path: str) -> list[dict]:
    """
    Procesa el archivo del cliente y retorna una lista de diccionarios con los hallazgos.
    Útil para integraciones vía API que requieren JSON en lugar de un archivo Excel.
    """
    if not os.path.exists(client_file_path):
        raise FileNotFoundError(f"Archivo de entrada no encontrado: {client_file_path}")
        
    ext = os.path.splitext(client_file_path)[1].lower()
    try:
        # Detectar magia de bytes para archivos Excel escondidos como CSVs
        is_excel = False
        with open(client_file_path, 'rb') as f:
            header = f.read(4)
            if header.startswith(b'PK\x03\x04') or header.startswith(b'\xd0\xcf\x11\xe0'):
                is_excel = True

        if is_excel or ext in ('.xlsx', '.xls'):
            df_client = pd.read_excel(client_file_path)
            # Detectar "CSV-in-Excel": un Excel con 1 sola columna cuyo header tiene comas
            # significa que alguien pegó texto CSV crudo dentro de Excel
            if len(df_client.columns) == 1 and ',' in str(df_client.columns[0]):
                logger.info("Detectado CSV-in-Excel: re-parseando como CSV...")
                from io import StringIO
                header_line = str(df_client.columns[0])
                data_lines = [str(v) for v in df_client.iloc[:, 0].tolist()]
                csv_text = header_line + '\n' + '\n'.join(data_lines)
                df_client = pd.read_csv(StringIO(csv_text), sep=',', on_bad_lines='skip')
        else:
            # Archivo de texto/CSV
            try:
                # 1. Autodetección de delimitador
                df_client = pd.read_csv(client_file_path, sep=None, engine='python', encoding='utf-8', on_bad_lines='skip')
            except Exception:
                try:
                    df_client = pd.read_csv(client_file_path, sep=None, engine='python', encoding='latin-1', on_bad_lines='skip')
                except Exception:
                    # 2. Fallback estricto a coma
                    df_client = pd.read_csv(client_file_path, sep=',', encoding='utf-8', on_bad_lines='skip')
    except Exception as e:
        logger.error(f"Error parseando archivo cliente: {e}")
        raise ValueError("Formato de archivo inválido.")

    # Validar columna NIT
    cols_upper = [str(c).upper().strip() for c in df_client.columns]
    nit_col = None
    
    # 1. Búsqueda estricta primero
    for idx, c in enumerate(cols_upper):
        if c == 'NIT' or c == 'IDENTIFICACION':
            nit_col = df_client.columns[idx]
            break
            
    # 2. Búsqueda laxa si no hay estricta
    if not nit_col:
        for idx, c in enumerate(cols_upper):
            if 'NIT' in c or 'IDENTIFICACION' in c or 'RUC' in c:
                nit_col = df_client.columns[idx]
                break
            
    if not nit_col:
        logger.warning("No se encontró columna NIT. Asumiendo la primera columna...")
        if len(df_client.columns) > 0:
            nit_col = df_client.columns[0]
        else:
            return [] # Archivo totalmente vacío
         
    def safe_normalize(x):
        if pd.isna(x):
            return ""
        try:
            return normalize_nit(x)
        except ValueError:
            return str(x).strip() # Retorna el valor crudo en vez de fallar
            
    df_client['NIT_NORMALIZADO'] = df_client[nit_col].apply(safe_normalize)
    # Ya no usamos dropna silencioso. Mantenemos todos los registros para entender qué falló.
    df_procesar = df_client.copy()
    
    nits_unicos = [n for n in df_procesar['NIT_NORMALIZADO'].unique().tolist() if str(n).isdigit()]
    df_bdme = consult_batch_bdme(nits_unicos) if nits_unicos else pd.DataFrame()
    
    resultados = []
    for _, row in df_procesar.iterrows():
        nit_raw = row[nit_col]
        nit_norm = row['NIT_NORMALIZADO']
        
        # Buscar nombre/razón social en columnas comunes
        name_cols = ['RAZON_SOCIAL', 'NOMBRE', 'RAZON SOCIAL', 'TERCERO', 'PROVEEDOR', 'DESCRIPCION']
        razon_social = str(nit_raw) # Default a mostrar lo que asumimos que es el NIT para debug
        for c in df_client.columns:
            if str(c).upper().strip() in name_cols:
                razon_social = row[c]
                break
        
        # Validar si el nit normalizado es válido para consultar
        if not nit_norm or not str(nit_norm).isdigit() or len(str(nit_norm)) < 6:
            resultados.append({
                "nit": str(nit_raw)[:20] if nit_raw else "VACÍO",
                "razon_social": str(razon_social).strip(),
                "en_ficticios": False,
                "estado_bdme": "INDETERMINADO",
                "facturador_habilitado": False,
                "nivel_riesgo": "REVISAR",
                "recomendacion": "Celda no es un NIT válido. Verifique el archivo."
            })
            continue

        en_ficticios = check_nit_dian(nit_norm)
        
        if not df_bdme.empty and 'nit' in df_bdme.columns:
            bdme_info = df_bdme[df_bdme['nit'] == nit_norm]
            estado_bdme = bdme_info['estado_bdme'].iloc[0] if not bdme_info.empty else "INDETERMINADO"
        else:
            estado_bdme = "INDETERMINADO"
        
        fact_habilitado = check_facturador_electronico(nit_norm)
        riesgo, recomendacion = calculate_risk(en_ficticios, estado_bdme, fact_habilitado)
        
        resultados.append({
            "nit": nit_norm,
            "razon_social": str(razon_social).strip(),
            "en_ficticios": en_ficticios,
            "estado_bdme": estado_bdme,
            "facturador_habilitado": fact_habilitado,
            "nivel_riesgo": riesgo,
            "recomendacion": recomendacion
        })

    return resultados

def generate_report(client_excel_path: str, output_path: str = None) -> str:
    """
    Genera el reporte final evaluando el riesgo de cada NIT dado.
    """
    resultados = get_audit_data(client_excel_path)
    df_resultado = pd.DataFrame(resultados)
    
    # Mapear nombres para compatibilidad con el resto del módulo
    df_resultado = df_resultado.rename(columns={
        "nit": "NIT",
        "razon_social": "RAZON_SOCIAL",
        "en_ficticios": "EN_FICTICIOS",
        "estado_bdme": "ESTADO_BDME",
        "facturador_habilitado": "FACTURADOR_HABILITADO",
        "nivel_riesgo": "NIVEL_RIESGO",
        "recomendacion": "RECOMENDACION"
    })
    
    # 5. Generar Excel
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    if not output_path:
        out_dir = os.path.join(os.path.dirname(__file__), '../outputs')
        os.makedirs(out_dir, exist_ok=True)
        output_path = os.path.join(out_dir, f"reporte_cumplimiento_{timestamp_str}.xlsx")
        
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Pestaña Detalle
            df_resultado.to_excel(writer, sheet_name='Detalle', index=False)
            
            # Pestaña Resumen
            total_analizados = len(df_resultado)
            if df_resultado.empty:
                # Columna mínima requerida para el resumen si está vacío
                counts = pd.Series(dtype=int)
            else:
                counts = df_resultado['NIVEL_RIESGO'].value_counts()
            
            resumen_data = {
                "Métrica": ["Total Analizados", "Total ALTO", "Total MEDIO", "Total REVISAR/INDETERMINADO", "Total BAJO", "Fecha Análisis"],
                "Valor": [
                    total_analizados,
                    counts.get("ALTO", 0),
                    counts.get("MEDIO", 0),
                    counts.get("REVISAR", 0),
                    counts.get("BAJO", 0),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ]
            }
            pd.DataFrame(resumen_data).to_excel(writer, sheet_name='Resumen', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Detalle']
            
            # Formatos condicionales de color
            from openpyxl.styles import PatternFill
            fill_rojo = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            fill_naranja = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
            fill_amarillo = PatternFill(start_color="FFFACC", end_color="FFFACC", fill_type="solid")
            fill_verde = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            
            # Nivel es G (columna 6) - 1-indexed in openpyxl -> 'F' level col is 6
            for row in range(2, worksheet.max_row + 1):
                cell_riesgo = worksheet.cell(row=row, column=6).value
                fill_color = None
                if cell_riesgo == 'ALTO':
                    fill_color = fill_rojo
                elif cell_riesgo == 'MEDIO':
                    fill_color = fill_naranja
                elif cell_riesgo == 'REVISAR':
                    fill_color = fill_amarillo
                elif cell_riesgo == 'BAJO':
                    fill_color = fill_verde
                
                if fill_color:
                     for col in range(1, worksheet.max_column + 1):
                         worksheet.cell(row=row, column=col).fill = fill_color
                         
        logger.info(f"Reporte exitosamente guardado en: {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"Error guardando Excel: {e}")
        raise

if __name__ == "__main__":
    # Test de Humo Report Engine
    # Crear un dummy excel client con NITs (Uno ficticio, otro no habilitado, otro de test)
    dummy_input = os.path.abspath(os.path.join(os.path.dirname(__file__), '../data/test_client_input.xlsx'))
    os.makedirs(os.path.dirname(dummy_input), exist_ok=True)
    pd.DataFrame({
        'NIT': ["900123456", "901565796", "801004345"], # El segundo esta en BDME, el tercero tambien
        'RAZON_SOCIAL': ["Empresa Dummy", "SERVICIOS & SUM", "COO SERVICIOS"]
    }).to_excel(dummy_input, index=False)
    
    print("Iniciando test de humo de report engine...")
    out_file = generate_report(dummy_input)
    print(f"✅ Test completado. Archivo en: {out_file}")
